from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any
import logging

class ExcelExportService:
    """Service for exporting messages to Excel format"""
    
    @staticmethod
    def create_messages_excel(messages: List[Dict[str, Any]], project_name: str) -> BytesIO:
        """
        Create an Excel file from a list of messages
        
        Args:
            messages: List of message dictionaries with username, generated_message, created_at
            project_name: Name of the project for the filename and header
            
        Returns:
            BytesIO: Excel file as bytes stream
        """
        try:
            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Messages"
            
            # Define colors and styles
            header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            data_font = Font(size=11)
            center_alignment = Alignment(horizontal="center", vertical="center")
            wrap_alignment = Alignment(wrap_text=True, vertical="top")
            
            # Add title row
            ws.merge_cells('A1:D1')
            title_cell = ws['A1']
            title_cell.value = f"DMify Messages Export - {project_name}"
            title_cell.font = Font(bold=True, size=14, color="4F46E5")
            title_cell.alignment = center_alignment
            
            # Add export info row
            ws.merge_cells('A2:D2')
            info_cell = ws['A2']
            info_cell.value = f"Exported on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            info_cell.font = Font(size=10, italic=True, color="666666")
            info_cell.alignment = center_alignment
            
            # Add empty row for spacing
            ws['A3'] = ""
            
            # Add headers
            headers = ["Username", "Generated Message", "Created Date", "Character Count"]
            header_row = 4
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Add data rows
            data_start_row = header_row + 1
            
            for row_idx, message in enumerate(messages, data_start_row):
                # Username (with @ prefix)
                username_cell = ws.cell(row=row_idx, column=1)
                username_cell.value = f"@{message['username']}"
                username_cell.font = data_font
                username_cell.alignment = center_alignment
                
                # Generated message
                message_cell = ws.cell(row=row_idx, column=2)
                message_cell.value = message['generated_message']
                message_cell.font = data_font
                message_cell.alignment = wrap_alignment
                
                # Created date (formatted)
                date_cell = ws.cell(row=row_idx, column=3)
                if isinstance(message['created_at'], str):
                    created_date = datetime.fromisoformat(message['created_at'].replace('Z', '+00:00'))
                else:
                    created_date = message['created_at']
                date_cell.value = created_date.strftime('%m/%d/%Y %I:%M %p')
                date_cell.font = data_font
                date_cell.alignment = center_alignment
                
                # Character count
                char_count_cell = ws.cell(row=row_idx, column=4)
                char_count_cell.value = len(message['generated_message'])
                char_count_cell.font = data_font
                char_count_cell.alignment = center_alignment
            
            # Auto-adjust column widths
            column_widths = {
                'A': 20,  # Username
                'B': 80,  # Message (wider for readability)
                'C': 20,  # Date
                'D': 15   # Character count
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Set row heights for better readability
            for row in range(data_start_row, data_start_row + len(messages)):
                ws.row_dimensions[row].height = 60  # Taller rows for wrapped text
            
            # Add footer with summary
            if messages:
                footer_row = data_start_row + len(messages) + 2
                ws.merge_cells(f'A{footer_row}:D{footer_row}')
                footer_cell = ws[f'A{footer_row}']
                footer_cell.value = f"Total Messages: {len(messages)} | Generated by DMify - AI Instagram DM Automation"
                footer_cell.font = Font(size=10, italic=True, color="666666")
                footer_cell.alignment = center_alignment
            
            # Save to BytesIO
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            logging.info(f"Successfully created Excel export with {len(messages)} messages for project: {project_name}")
            return excel_buffer
            
        except Exception as e:
            logging.error(f"Error creating Excel export: {str(e)}")
            raise Exception(f"Failed to create Excel file: {str(e)}")
    
    @staticmethod
    def get_filename(project_name: str) -> str:
        """
        Generate a safe filename for the Excel export
        
        Args:
            project_name: Name of the project
            
        Returns:
            str: Safe filename with timestamp
        """
        # Clean project name for filename
        safe_project_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).strip()
        safe_project_name = safe_project_name.replace(' ', '_')
        
        # Add timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        return f"DMify_Messages_{safe_project_name}_{timestamp}.xlsx"
    
    @staticmethod
    def validate_export_eligibility(user_subscription: Dict[str, Any]) -> bool:
        """
        Check if user is eligible for Excel export (Tier 2+ only)
        
        Args:
            user_subscription: User's subscription data
            
        Returns:
            bool: True if user can export, False otherwise
        """
        if not user_subscription:
            return False
        
        plan_id = user_subscription.get('plan_id')
        eligible_plans = ['plan_2', 'plan_3']  # Growth and Pro plans
        
        return plan_id in eligible_plans
